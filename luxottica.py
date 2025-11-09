from io import BytesIO
import os
import sys
import json
import glob
import requests
import threading
from PIL import Image
from time import sleep
from openpyxl import Workbook
from datetime import datetime
from selenium import webdriver
from models.store import Store
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
from selenium.webdriver.common.by import By
from openpyxl.drawing.image import Image as Imag
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service as ChromeService


class myScrapingThread(threading.Thread):
    def __init__(self, thread_id, name, scraper, varinat: dict, brand: str, glasses_type: str, headers: dict, tokenValue: str):
        threading.Thread.__init__(self)
        self.thread_id = thread_id
        self.name = name
        self.scraper = scraper
        self.varinat = varinat
        self.brand = brand
        self.glasses_type = glasses_type
        self.headers = headers
        self.tokenValue = tokenValue
        self.status = "in progress"

    def run(self):
        try:
            self.scraper.get_variants(self.varinat, self.brand, self.glasses_type, self.headers, self.tokenValue)
            self.status = "completed"
        except Exception as e:
            self.status = "failed"
            self.scraper.print_logs(f'Exception in thread {self.name}: {e}')

class myScrapingThreadController:
    def __init__(self, max_threads: int, DEBUG, logs_filename: str) -> None:
        self.thread_list = []
        self.thread_counter = 0
        self.max_threads = max_threads
        self.lock = threading.Lock()
        self.data = []
        self.DEBUG = DEBUG
        self.logs_filename = logs_filename

    def create_thread(self, varinat: dict, brand: str, glasses_type: str, headers: dict, tokenValue: str):
        self.manage_thread_list()
        thread_name = "Thread-" + str(self.thread_counter)
        new_thread = myScrapingThread(self.thread_counter, thread_name, self, varinat, brand, glasses_type, headers, tokenValue)
        
        with self.lock:
            self.thread_list.append(new_thread)
            self.thread_counter += 1
        
        new_thread.start()

    def clean_completed_threads(self):
        with self.lock:
            self.thread_list = [thread for thread in self.thread_list if thread.status != "completed"]

    def manage_thread_list(self):
        while len(self.thread_list) >= self.max_threads:
            self.clean_completed_threads()
            sleep(1)

    def is_thread_list_completed(self) -> bool:
        with self.lock:
            for obj in self.thread_list:
                if obj.status == "in progress":
                    return False
            return True

    def wait_for_thread_list_to_complete(self):
        while not self.is_thread_list_completed():
            sleep(1)
        with self.lock:
            self.thread_counter = 0
            self.thread_list.clear()

    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total:
            print()

    def get_variants(self, varinat: dict, brand: str, glasses_type: str, headers: dict, tokenValue: str):
        try:
            product = Product()
            product.brand = brand
            product.url = f'https://my.essilorluxottica.com/myl-it/en-GB/pdp/{str(varinat["partNumber"]).replace(" ", "+").replace("_", "-").replace("/", "-").lower()}'
            product.number = str(varinat['partNumber']).strip().split('_')[0].strip()[1:]
            product.name = str(varinat['name']).strip()
            product.frame_code = str(varinat['partNumber']).strip().split('_')[-1].strip()
            product.status = 'active'
            product.type = str(glasses_type).strip().title()

            prices = self.get_prices(varinat['uniqueID'], headers)
            
            metafields = Metafields()
           
            properties = self.get_product_variants(varinat['uniqueID'], headers)
            product.frame_color = properties['frame_color'] if 'frame_color' in properties else ''
            product.lens_color = properties['lens_color'] if 'lens_color' in properties else ''
            metafields.for_who = properties['for_who'] if 'for_who' in properties else ''
            metafields.lens_material = properties['lens_material'] if 'lens_material' in properties else ''
            metafields.frame_shape = properties['frame_shape'] if 'frame_shape' in properties else ''
            metafields.frame_material = properties['frame_material'] if 'frame_material' in properties else ''
            metafields.lens_technology = properties['lens_technology'] if 'lens_technology' in properties else ''
            metafields.img_url = properties['img_url'] if 'img_url' in properties else ''
            
            
            sizes = properties['sizes'] if 'sizes' in properties else []

            barcodes, product_sizes = [], []
            for size in sizes:
                variant = Variant()
                variant.title = size['title']
                variant.sku = f'{product.number} {product.frame_code} {variant.title}'
                variant.inventory_status = size['inventory_status']
                variant.found_status = 1
                variant.barcode_or_gtin = size['UPC']
                barcodes.append(size['UPC'])
                variant.weight = '0.5'
                variant.size = size['size']
                product_sizes.append(size['size'])
                for price in prices:
                    variant.wholesale_price = price['wholesale_price']
                    variant.listing_price = price['listing_price']
                product.variants = variant

            metafields.product_size = ', '.join(product_sizes)
            metafields.gtin1 = ', '.join(barcodes)

            # images_360 = self.get_360_images(tokenValue, headers)
            # if images_360:
            #     for image360 in images_360:
            #         if image360 not in metafields.img_360_urls:
            #             metafields.img_360_urls = image360
            
            
            product.metafields = metafields
            
            self.data.append(product)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_variants: {e}')
            self.print_logs(f'Exception in get_variants: {e}')

    def get_prices(self, tokenValue: str, headers: dict) -> list[str]:
        prices = []
        url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/prices?productId={tokenValue}'
        for _ in range(0, 3):
            try:
                response = requests.get(url=url, headers= headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    for data in json_data['data']:
                        wholesale_price, listing_price = '', ''
                        try: wholesale_price = float(data[tokenValue]['OPT'][0]['price']['value'])
                        except: pass
                        try: listing_price = float(data[tokenValue]['PUB'][0]['price']['value'])
                        except: pass
                        prices.append({'wholesale_price': wholesale_price, 'listing_price': listing_price})
            except Exception as e:
                if self.DEBUG: print(f'Exception in get_prices: {e}: url: {url}')
                self.print_logs(f'Exception in get_prices: {e}: url: {url}')
            if prices: break
            else: sleep(0.5)
        return prices

    def get_product_variants(self, uniqueID: str, headers: dict) -> dict:
        properties = {}
        url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/variants/{uniqueID}'
        for _ in range(0, 3):
            try:
                response = requests.get(url=url, headers= headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    frame_color, lens_color, for_who, lens_material, frame_shape, frame_material, lens_technology = '', '', '', '', '', '', ''
                    img_url = f"{str(json_data['data']['catalogEntryView'][0]['fullImage']).strip()}?impolicy=MYL_EYE&wid=600"
                    for attribute in json_data['data']['catalogEntryView'][0]['attributes']:
                        values = []
                        if attribute['identifier'] == 'FRONT_COLOR_DESCRIPTION':
                            for value in attribute['values']: values.append(value['value'])
                            frame_color = ', '.join(values)
                        elif attribute['identifier'] == 'LENS_COLOR_DESCRIPTION':
                            for value in attribute['values']: values.append(value['value'])
                            lens_color = ', '.join(values)
                        elif attribute['identifier'] == 'GENDER':
                            for value in attribute['values']: values.append(value['value'])
                            for_who = ', '.join(values)
                        elif attribute['identifier'] == 'LENS_MATERIAL':
                            for value in attribute['values']: values.append(value['value'])
                            lens_material = ', '.join(values)
                        elif attribute['identifier'] == 'FACE_SHAPE':
                            for value in attribute['values']: values.append(value['value'])
                            frame_shape = ', '.join(values)
                        elif attribute['identifier'] == 'FRAME_MATERIAL':
                            for value in attribute['values']: values.append(value['value'])
                            frame_material = ', '.join(values)
                        elif attribute['identifier'] == 'PHOTOCHROMIC':
                            if attribute['values'][0]['value'] == 'TRUE':
                                if lens_technology: lens_technology += str(' PHOTOCHROMIC').title()
                                else: lens_technology = str('PHOTOCHROMIC').title()
                        elif attribute['identifier'] == 'POLARIZED':
                            if attribute['values'][0]['value'] == 'TRUE':
                                if lens_technology: lens_technology += str(' POLARIZED').title()
                                else: lens_technology = str('POLARIZED').title()

                    if not str(lens_technology).strip():
                        for attribute in json_data['data']['catalogEntryView'][0]['attributes']:
                            if attribute['identifier'] == 'LENS_COLORING_PERCEIVED':
                                lens_technology = str(attribute['values'][0]['value']).strip()

                    ids = []
                    sizes_without_q = []
                    for sKU in json_data['data']['catalogEntryView'][0]['sKUs']:
                        BRIDGE, SIZE, TEMPLE = '', '', ''
                        uniqueID = str(sKU['uniqueID'])
                        title = str(sKU['partNumber']).strip()[-2:]
                        upc = str(sKU['upc'])
                        ids.append(uniqueID)
                        for attribute in sKU['attributes']:
                            if attribute['identifier'] == 'BRIDGE_WIDTH':
                                BRIDGE = attribute['values'][0]['value']
                            elif attribute['identifier'] == 'FRAME_SIZE':
                                SIZE = attribute['values'][0]['value']
                            elif attribute['identifier'] == 'TEMPLE_LENGTH':
                                TEMPLE = attribute['values'][0]['value']

                        sizes_without_q.append({'uniqueID': uniqueID, 'title': title, 'UPC': upc, 'size': f'{BRIDGE}-{SIZE}-{TEMPLE}'})

                    sizes = []
                    json_response = self.check_availability('%2C'.join(ids), headers)
                    for json_res in json_response:
                        productId = json_res['productId']
                        for size_without_q in sizes_without_q:
                            if productId == size_without_q['uniqueID']:
                                inventory_status = str(json_res['x_state']).strip()
                                # if str(json_res['x_state']).strip().upper() == 'AVAILABLE': inventory_quantity = 1

                                sizes.append({'title': size_without_q['title'], 'inventory_status': inventory_status, "UPC": size_without_q['UPC'], "size": size_without_q['size']})
                    
                    properties = {
                        'img_url': img_url if img_url else '',
                        'frame_color': frame_color if frame_color else '', 
                        'lens_color': lens_color if lens_color else '', 
                        'for_who': for_who if for_who else '', 
                        'lens_material': lens_material if lens_material else '', 
                        'frame_shape': frame_shape if frame_shape else '', 
                        'frame_material': frame_material if frame_material else '', 
                        'lens_technology': lens_technology if lens_technology else '',
                        'sizes': sizes if sizes else []
                    }
                else: self.print_logs(f'Status code: {response.status_code} for id and tokenValue')
            except Exception as e:
                if self.DEBUG: print(f'Exception in get_tokenValue_and_id: {e}: url: {url}')
                self.print_logs(f'Exception in get_tokenValue_and_id: {e}: url: {url}')
            if properties: break
            else: sleep(0.5)
        
        return properties

    def check_availability(self, payload: str, headers: dict) -> list[dict]:
        json_data = {}
        url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/availability?productId={payload}'
        for _ in range(0, 3):
            try:
                response = requests.get(url=url, headers=headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    json_data = json_data['data']
                    json_data = json_data['doorInventoryAvailability'][0]['inventoryAvailability']
            except Exception as e:
                if self.DEBUG: print(f'Exception in check_availability: {e} : url: {url}')
                self.print_logs(f'Exception in check_availability: {e} : url: {url}')
            if json_data: break
            else: sleep(0.5)
        return json_data

    def get_360_images(self, tokenValue: str, headers: dict) -> list[str]:
        image_360_urls = []
        try:
            counter = 0
            while True:
                counter += 1
                url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/variants/{tokenValue}/attachments?type=PHOTO_360'
                response = requests.get(url=url, headers= headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    for attachment in json_data['data']['catalogEntryView'][0]['attachments']:
                        image_360_urls.append(attachment['attachmentAssetPath'])
                    break
                if counter == 3: break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_360_images: {e}')
            self.print_logs(f'Exception in get_360_images: {e}')
        finally: return image_360_urls

    def get_images(self, tokenValue: str, headers: dict) -> list[str]:
        images = []
        try:
            url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/variants/{tokenValue}/attachments?type=PHOTO'
            response = requests.get(url=url, headers= headers)
            if response.status_code == 200:
                json_data = json.loads(response.text)
                for attachment in json_data['data']['catalogEntryView'][0]['attachments']:
                    images.append(f"{attachment['attachmentAssetPath']}?impolicy=MYL_EYE&wid=834")
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_images: {e}')
            self.print_logs(f'Exception in get_images: {e}')
        finally: return images


class Luxottica_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str, chrome_path: str) -> None:
        self.DEBUG = DEBUG
        self.data = []
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.max_threads = 15
        self.myScrapingThreadController_obj = myScrapingThreadController(self.max_threads, self.DEBUG, self.logs_filename)
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.browser = webdriver.Chrome(service=ChromeService(chrome_path), options=self.chrome_options)
        pass

    def controller(self, store: Store, brands_with_types: list[dict]) -> None:
        try:
            self.initialize_browser(store.link)
            
            for brand_with_type in brands_with_types:
                login_flag = False
                if '/login' in self.browser.current_url:
                    if self.login(store.link, store.username, store.password):
                        sleep(10)
                        login_flag = True
                    else: 
                        print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
                        self.print_logs(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
                else: login_flag = True

                if login_flag:
                    cookies, dtPC = '', ''
                    brand: str = brand_with_type['name']
                    brand_url_json = brand_with_type['url']
                    glasses_types = brand_with_type['glasses_type']

                    print(f'Brand: {brand}')
                    self.print_logs(f'Brand: {brand}')

                    for glasses_type_index, glasses_type in enumerate(glasses_types):
                        
                        brand_url = self.select_category(brand_url_json, glasses_type, store.username, store.password)
                        if brand_url:
                            total_products = self.get_total_products_for_brand()
                        
                            print(f'Total products found: {total_products} | Type: {glasses_type}')
                            self.print_logs(f'Total products found: {total_products} | Type: {glasses_type}')

                            if int(total_products) > 0:
                                page_number = 1
                                scraped_products = 0
                                start_time = datetime.now()
                                print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                self.print_logs(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                
                                self.myScrapingThreadController_obj.printProgressBar(0, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                while int(scraped_products) != int(total_products):
                                    product_divs = self.get_product_divs_on_page()
                                    counter = 0
                                    while counter < len(product_divs):
                                        
                                        scraped_products += 1
                                        try:    
                                            url = str(product_divs[counter].find_element(By.CSS_SELECTOR, 'a[class^="Tile__ImageContainer"]').get_attribute('href'))
                                            identifier = str(url).split('/')[-1].strip()

                                            if not cookies: cookies = self.get_cookies_from_browser(identifier)

                                            headers = self.get_headers(cookies, url, dtPC)
                                            tokenValue = self.get_tokenValue(identifier, headers)
                                            if tokenValue: 
                                                parentCatalogEntryID = self.get_parentCatalogEntryID(tokenValue, headers)
                                                if parentCatalogEntryID: 
                                                    variants = self.get_all_variants_data(parentCatalogEntryID, headers)
                                                    for variant in variants:
                                                        self.myScrapingThreadController_obj.create_thread(variant, brand,  glasses_type, headers, tokenValue)
                                                
                                        except Exception as e:
                                            if self.DEBUG: print(f'Exception in loop: {e}')
                                            self.print_logs(f'Exception in loop: {e}')

                                        self.myScrapingThreadController_obj.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                        counter += 1    
                                    if int(scraped_products) < int(total_products):
                                        page_number += 1
                                        self.move_to_next_page(brand_url, page_number)
                                        self.wait_until_element_found(40, 'css_selector', 'div[class^="PLPTitle__Section"] > p[class^="CustomText__Text"]')
                                        total_products = self.get_total_products_for_brand()
                                    else: break
                                
                                self.myScrapingThreadController_obj.wait_for_thread_list_to_complete()
                                self.save_to_json(self.myScrapingThreadController_obj.data)

                                end_time = datetime.now()
                                
                                print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                print('Duration: {}\n'.format(end_time - start_time))
                                self.print_logs('Duration: {}\n'.format(end_time - start_time))

                        else: 
                            print(f'Cannot find {glasses_type} for {brand}')
                            self.print_logs(f'Cannot find {glasses_type} for {brand}')
                
                self.browser.get('https://my.essilorluxottica.com/myl-it/en-GB/homepage')
                self.wait_until_browsing()
        except Exception as e: self.print_logs(f'Exception in Luxottica_All_Scraper controller: {e}')
        finally: self.browser.quit()

    # initialize the browser
    def initialize_browser(self, url: str) -> None:
        try:
            self.browser.get(url)
            self.wait_until_browsing()
        except Exception as e: self.print_logs(f'Exception in initialize_browser: {e}')

    # browser function to wait until the page is completely loaded
    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    # logging function
    def login(self, url, username: str, password: str) -> bool:
        login_flag = False
        while not login_flag:
            try:
                self.accept_cookies_before_login()
                if self.wait_until_element_found(10, 'xpath', '//input[@id="signInName"]'):
                    for _ in range(0, 30):
                        try:
                            self.browser.find_element(By.XPATH, '//input[@id="signInName"]').send_keys(username)
                            break
                        except: sleep(0.3)
                    sleep(0.2)
                    
                    if self.wait_until_element_found(20, 'xpath', '//button[@id="continue"]'):
                        for _ in range(0, 30):
                            try:
                                self.browser.find_element(By.XPATH, '//button[@id="continue"]').click()
                                break
                            except: sleep(0.3)

                        if self.wait_until_element_found(20, 'xpath', '//input[@id="password"]'):
                            for _ in range(0, 30):
                                try:
                                    self.browser.find_element(By.XPATH, '//input[@id="password"]').send_keys(password)
                                    break
                                except: sleep(0.5)
                            sleep(0.2)
                            self.browser.find_element(By.XPATH, '//button[@id="next"]').click()
                            self.wait_until_browsing()
                            for _ in range(0, 100):
                                try:
                                    brand_xpath_value = ''
                                    if '/myl-it/it-IT/homepage' in self.browser.current_url: brand_xpath_value = "//button[contains(text(), 'Brand')]"
                                    elif '/myl-it/en-GB/homepage' in self.browser.current_url: brand_xpath_value = "//button[contains(text(), 'Brands')]"
                                    a = self.browser.find_element(By.XPATH, brand_xpath_value)
                                    if a: 
                                        login_flag = True
                                        if '/homepage' in self.browser.current_url:
                                            self.browser.get('https://my.essilorluxottica.com/myl-it/en-GB/homepage')
                                        self.accept_cookies_after_login()
                                        break
                                    else: sleep(0.3)
                                except: sleep(0.3)
                        else: print('Password input not found')
                else: print('Email input not found')
            except Exception as e:
                self.print_logs(f'Exception in login: {str(e)}')
                if self.DEBUG: print(f'Exception in login: {str(e)}')

            if not login_flag: self.initialize_browser(url)
        return login_flag

    # function to accept all the cookies before starting the login process
    def accept_cookies_before_login(self) -> None:
        try:
            if self.wait_until_element_found(5, 'css_selector', 'div[class^="CookiesBanner__SecondButtonWrap"] > button'):
                self.browser.find_element(By.CSS_SELECTOR, 'div[class^="CookiesBanner__SecondButtonWrap"] > button').click()
                sleep(0.3)
        except Exception as e:
            self.print_logs(f'Exception in accept_cookies_before_login: {str(e)}')
            if self.DEBUG: print(f'Exception in accept_cookies_before_login: {str(e)}')

    # function to wait untill the object is loaded to the page
    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    # function to accept cookies after the login process is completed    
    def accept_cookies_after_login(self) -> None:
        try:
            if self.wait_until_element_found(5, 'css_selector', 'div[class^="CookiesContent__Container"] > div > button[class$="underline"]'):
                btn = self.browser.find_element(By.CSS_SELECTOR, 'div[class^="CookiesContent__Container"] > div > button[class$="underline"]')
                ActionChains(self.browser).move_to_element(btn).click().perform()
                sleep(0.3)
        except Exception as e:
            self.print_logs(f'Exception in accept_cookies_after_login: {str(e)}')
            if self.DEBUG: print(f'Exception in accept_cookies_after_login: {str(e)}')

    # function to select the category of brand        
    def select_category(self, url: str, glasses_type: str, username: str, password: str) -> str:
        brand_url = ''
        for _ in range(0, 10):
            try:
                if url:
                    self.wait_until_element_found(30, 'xpath', "//span/button[contains(text(),'Brands')]")
                    ActionChains(self.browser).move_to_element(self.browser.find_element(By.XPATH, "//span/button[contains(text(),'Brands')]")).perform()
                    sleep(0.5)
                    self.browser.get(url)
                    self.wait_until_browsing()
                    sleep(5)

                    if self.browser.current_url == url:
                        category_css_selector = ''
                        if glasses_type == 'Sunglasses': category_css_selector = 'button[data-element-id^="Categories_sunglasses_"]'
                        elif glasses_type == 'Sunglasses Kids': category_css_selector = 'button[data-element-id^="Categories_sunglasses-kids"]'
                        elif glasses_type == 'Eyeglasses': category_css_selector = 'button[data-element-id^="Categories_eyeglasses_"]'
                        elif glasses_type == 'Eyeglasses Kids': category_css_selector = 'button[data-element-id^="Categories_eyeglasses-kids"]'
                        elif glasses_type == 'Goggles and helmets': category_css_selector = 'button[data-element-id^="Categories_adult_ViewAll"]'#'button[data-element-id^="Categories_gogglesHelmets"]'
                        elif glasses_type == 'Goggles and helmets kids': category_css_selector = 'button[data-element-id^="Categories_children_ViewAll"]'

                        if self.wait_until_element_found(20, 'css_selector', category_css_selector):
                            element = self.browser.find_element(By.CSS_SELECTOR, category_css_selector)
                            ActionChains(self.browser).move_to_element(element).perform()
                            sleep(0.5)
                            ActionChains(self.browser).move_to_element(element).click().perform()
                            sleep(0.4)

                            for _ in range(0, 100):
                                try:
                                    value = str(self.browser.find_element(By.CSS_SELECTOR, 'div[class^="PLPTitle__Section"] > p[class^="CustomText__Text"]').text).strip()
                                    if '(' in value or ')' in value: break
                                except: sleep(0.5)

                            brand_url = self.browser.current_url
                            break
                        else: break
                else: break
            except Exception as e:
                if self.DEBUG: print(f'Exception in select_category: {e}')
                self.print_logs(f'Exception in select_category: {e}')
        return brand_url

    # function to get total number of products for the brand
    def get_total_products_for_brand(self) -> int:
        total_products = 0
        try:
            for _ in range(0, 200):
                try:
                    total_sunglasses = str(self.browser.find_element(By.CSS_SELECTOR, 'div[class^="PLPTitle__Section"] > p[class^="CustomText__Text"]').text).strip()
                    if '(' in total_sunglasses:
                        total_sunglasses = total_sunglasses.split('(')[-1].strip().replace(')', '').strip()
                        if total_sunglasses: total_products = int(total_sunglasses)
                        else: total_products = 0
                        break
                    else: sleep(0.3)
                except: 
                    try:
                        text = str(self.browser.find_element(By.CSS_SELECTOR, 'div[class^="PLPGeneric__MainColumn"] > div > p').text).strip()
                        if 'Sorry, there are no products' in text: break
                        else: sleep(0.3)
                    except: sleep(0.3)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_total_products_and_pages: {str(e)}')
            self.print_logs(f'Exception in get_total_products_and_pages: {str(e)}')
        finally: return total_products

    # get all the products on the page
    def get_product_divs_on_page(self) -> list:
        product_divs = []
        for _ in range(0, 30):
            try:
                product_divs = self.browser.find_elements(By.CSS_SELECTOR, 'div[data-element-id="Tiles"] > div[class^="Tile"]')
                for product_div in product_divs:
                    product_number = str(product_div.get_attribute('data-description')).strip()
                    product_name = str(product_div.find_element(By.CSS_SELECTOR, 'div[class^="TileHeader__Header"] > div > span').text).strip()
                    total_varinats_for_product = str(product_div.find_element(By.CSS_SELECTOR, 'div[class^="Tile__ColorSizeContainer"] > div > span').text).strip()

                break
            except: sleep(0.5)
        return product_divs
    
    # function to get cookies from the browser by using provided identifier
    # this function is used to get the cookies for the product page
    def get_cookies_from_browser(self, identifier: str) -> str:
        cookies = ''
        try:
            self.open_new_tab(f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/pages/identifier/{identifier}')
            sleep(2)
            browser_cookies = self.browser.get_cookies()
        
            for browser_cookie in browser_cookies:
                if browser_cookie['name'] == 'dtPC': dtPC = browser_cookie['value']
                cookies = f'{browser_cookie["name"]}={browser_cookie["value"]}; {cookies}'
            cookies = cookies.strip()[:-1]
            self.close_last_tab()
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_cookies_from_browser: {e}')
            self.print_logs(f'Exception in get_cookies_from_browser: {e}')
        finally: return cookies

    # function to open the provided url in new tab of browser
    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()
    
    # function to close the last tab of the browser
    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])

    # function to get the headers for the request
    def get_headers(self, cookie: str, referer: str, dtpc: str) -> dict:
        return {
            'accept': 'application/json, text/plain, */*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.9',
            'cookie': cookie,
            'referer': referer,
            'sec-ch-ua': '"Chromium";v="106", "Google Chrome";v="106", "Not;A=Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36',
            'x-dtpc': dtpc
        }

    # function to get the token value for the product
    def get_tokenValue(self, identifier: str, headers: dict):
        tokenValue = ''
        url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/pages/identifier/{identifier}'
        for _ in range(0, 3):
            try:
                response = requests.get(url=url, headers=headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    id = str(int(json_data['data']['contents'][0]['id']))
                    tokenValue = json_data['data']['contents'][0]['tokenValue']
                else: self.print_logs(f'Status code: {response.status_code} for id and tokenValue {response.text} {response.headers}')
            except Exception as e:
                if self.DEBUG: print(f'Exception in get_tokenValue: {e}, url: {url}')
                self.print_logs(f'Exception in get_tokenValue: {e}, url: {url}')
            if tokenValue: break
            else: sleep(0.5)
        return tokenValue

    # function to get parent catalog entry id for the product
    def get_parentCatalogEntryID(self, tokenValue: str, headers: dict) -> str:
        parentCatalogEntryID = ''
        url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/variants/{tokenValue}'
        for _ in range(0, 3):
            try:
                response = requests.get(url=url, headers=headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    parentCatalogEntryID = json_data['data']['catalogEntryView'][0]['parentCatalogEntryID']
                else: self.print_logs(f'Status code: {response.status_code} for parentCatalogEntryID')
            except Exception as e:
                if self.DEBUG: print(f'Exception in get_parentCatalogEntryID: {e}, url: {url}')
                self.print_logs(f'Exception in get_parentCatalogEntryID: {e}, url: {url}')
            if parentCatalogEntryID: break
            else: sleep(0.5)
        return parentCatalogEntryID

    # function to get all the variants data for the product
    def get_all_variants_data(self, parentCatalogEntryID: str, headers: str) -> list[dict]:
        variants = []
        url = f'https://my.essilorluxottica.com/fo-bff/api/priv/v1/myl-it/en-GB/products/{parentCatalogEntryID}/variants'
        for _ in range(0, 3):
            try:
                response = requests.get(url=url, headers=headers)
                if response.status_code == 200:
                    json_data = json.loads(response.text)
                    for index, variant in enumerate(json_data['data']['catalogEntryView'][0]['variants']):
                        try:
                            name = ''
                            partNumber = variant['partNumber']
                            uniqueID = variant['uniqueID']
                            try: name = variant['name']
                            except: pass
                          

                            variants.append({'sequence': (index+1), 'partNumber': partNumber, 'name': name, 'uniqueID': uniqueID})
                        except: pass
                else: self.print_logs(f'Status code: {response.status_code} for get_all_variants_data')
            except Exception as e:
                if self.DEBUG: print(f'Exception in get_all_variants_data: {e}')
                self.print_logs(f'Exception in get_all_variants_data: {e}')
            if variant: break
            else: sleep(0.5)
        return variants

    # function to move to the next variants grid
    def move_to_next_varinats_grid(self) -> None:
        for _ in range(0, 30):
            try:
                variants_divs = self.browser.find_elements(By.CSS_SELECTOR, 'div[data-element-id="Variants"]')
                btn = variants_divs[1].find_elements(By.CSS_SELECTOR, 'div[class^="IconButton__Container"] > button')[1]
                if 'button-out-of-stock' in btn.get_attribute('class'): break
                else: 
                    btn.click()
                    for _ in range(0, 30):
                        try:
                            variants_divs = self.browser.find_elements(By.CSS_SELECTOR, 'div[data-element-id="Variants"]')
                            if len(variants_divs) == 2:
                                variants_grids = variants_divs[1].find_elements(By.CSS_SELECTOR, 'div[class^="ExpandedTile__TilesSection"] > div')
                                if len(variants_grids) > 0: break
                        except: sleep(0.3)
                    break
            except: pass

    # function to move to the next page of the brand
    def move_to_next_page(self, brand_url: str, page_number: int) -> None:
        self.browser.get(f'{brand_url}&pageNumber={page_number}')
        self.wait_until_browsing()
        sleep(0.8)

    # functiion to save the data to json file
    def save_to_json(self, products: list[Product]):
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_status': variant.inventory_status,
                        'found_status': variant.found_status,
                        'listing_price': variant.listing_price,
                        'wholesale_price': variant.wholesale_price,
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                        'weight': variant.weight
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code, 
                    'frame_color': product.frame_color, 
                    'lens_code': product.lens_code, 
                    'lens_color': product.lens_color, 
                    'status': product.status, 
                    'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.metafields.img_url },
                        { 'key': 'img_360_urls', 'value': product.metafields.img_360_urls }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)

            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            self.print_logs(f'Exception in save_to_json: {e}')

    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass


def read_file(filename: str):
    f = open(filename)
    data = f.read()
    f.close()
    return data

def read_data_from_json_file(DEBUG, result_filename: str, logs_filename: str) -> list:
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color = '', '', '', '', '', ''
                brand = json_d['brand']
                number = str(json_d['number']).strip().upper()
                if '/' in number: number = number.replace('/', '-').strip()
                frame_code = str(json_d['frame_code']).strip().upper()
                if '/' in frame_code: frame_code = frame_code.replace('/', '-').strip()
                frame_color = str(json_d['frame_color']).strip().title()
                lens_color = str(json_d['lens_color']).strip().title()
                
                for json_metafiels in json_d['metafields']:
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                for json_variant in json_d['variants']:
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    inventory_status = json_variant['inventory_status']
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()
                    barcode_or_gtin = str(json_variant['barcode_or_gtin']).strip()
                    img_url = img_url.replace('impolicy=MYL_EYE&wid=262', 'impolicy=MYL_EYE&wid=600')
                    image_filename = f'Images/{sku.replace("/", "_")}.jpg'
                    if not os.path.exists(image_filename):
                        download_and_save_image(img_url, logs_filename, image_filename)
                       
                    data.append([number, frame_code, frame_color, lens_color, brand, sku, wholesale_price, listing_price, barcode_or_gtin, inventory_status, image_filename])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_and_save_image(url: str, logs_filename: str, image_filename: str) -> None:
    try:
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=60)
                response.raise_for_status()

                image = Image.open(BytesIO(response.content))
                image.verify()

                image = Image.open(BytesIO(response.content)).convert("RGB")
                image.save(image_filename, "JPEG")

                break
            except Exception as e: 
                print_logs(f'Exception in download_image request: {str(e)} for {url}', logs_filename)
                sleep(0.3)
            counter += 1
            if counter == 10:
                print_logs(f'Max retries reached for {url}', logs_filename)
                break
    except Exception as e: print_logs(f'Exception in download_image: {str(e)}')


def image_download_process(image_filename, image_url):
    try:  
        print('Downloading image: ', image_filename)
        image_attachment = download_image(image_url)
        if image_attachment:
            with open(image_filename, 'wb') as f: f.write(image_attachment)
    except Exception as e:
        print(f'Exception in image_download_process: {str(e)} for {image_url}')

def saving_picture_in_excel(data: list, excel_results_filename: str):
    try:
        workbook = Workbook()
        worksheet = workbook.active

        worksheet.cell(row=1, column=1, value='Model Code')
        worksheet.cell(row=1, column=2, value='Lens Code')
        worksheet.cell(row=1, column=3, value='Color Frame')
        worksheet.cell(row=1, column=4, value='Color Lens')
        worksheet.cell(row=1, column=5, value='Brand')
        worksheet.cell(row=1, column=6, value='SKU')
        worksheet.cell(row=1, column=7, value='Wholesale Price')
        worksheet.cell(row=1, column=8, value='Listing Price')
        worksheet.cell(row=1, column=9, value='UPC')
        worksheet.cell(row=1, column=10, value='Item Status')
        worksheet.cell(row=1, column=11, value="Image")

        for index, d in enumerate(data):
            new_index = index + 2
            worksheet.cell(row=new_index, column=1, value=d[0])
            worksheet.cell(row=new_index, column=2, value=d[1])
            worksheet.cell(row=new_index, column=3, value=d[2])
            worksheet.cell(row=new_index, column=4, value=d[3])
            worksheet.cell(row=new_index, column=5, value=d[4])
            worksheet.cell(row=new_index, column=6, value=d[5])
            worksheet.cell(row=new_index, column=7, value=d[6])
            worksheet.cell(row=new_index, column=8, value=d[7])
            worksheet.cell(row=new_index, column=9, value=d[8])
            worksheet.cell(row=new_index, column=10, value=d[9])

            image_filename = f'Images/{d[5].replace("/", "_")}.jpg'
            

            if os.path.exists(image_filename):
                try:
                    im = Image.open(image_filename)
                        
                    width, height = im.size
                    worksheet.row_dimensions[new_index].height = height

                    worksheet.add_image(Imag(image_filename), anchor='K'+str(new_index))
                except Exception as e:
                    print(f'Exception in adding image to excel: {str(e)} for {image_filename}') 
                    input('wait')
                    pass
                    
        workbook.save(excel_results_filename)
    except Exception as e: print(f'Exception in saving_picture_in_excel: {str(e)}')   

# print logs to the log file
def print_logs(log: str, logs_filename) -> None:
    try:
        with open(logs_filename, 'a') as f:
            f.write(f'\n{log}')
    except: pass

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False

    chrome_path = ChromeDriverManager().install()
    if 'chromedriver.exe' not in chrome_path:
        chrome_path = str(chrome_path).split('/')[0].strip()
        chrome_path = f'{chrome_path}\\chromedriver.exe'

    # create directories
    if not os.path.exists('requirements'): os.makedirs('requirements')
    if not os.path.exists('Logs'): os.makedirs('Logs')
    if not os.path.exists('Images'): os.makedirs('Images')
    
    start_json_filename = 'Start.json'
    credentails_filename = 'requirements/credentails.json'
    json_results_filename = 'requirements/json_results.json'
    excel_results_filename = 'Results.xlsx'
    
    # remove old files
    if os.path.exists(json_results_filename): os.remove(json_results_filename)
    if os.path.exists(excel_results_filename): os.remove(excel_results_filename)
    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')
    for filename in glob.glob('Images/*'): os.remove(filename)


    if os.path.exists(start_json_filename):
        json_data = json.loads(read_file(start_json_filename))
        
        if 'brands' in json_data:
            brands = json_data.get('brands')

            if os.path.exists(credentails_filename):
                credentails_json_data = json.loads(read_file(credentails_filename))

                store = Store()
                store.link = credentails_json_data.get('url')
                store.username = credentails_json_data.get('username')
                store.password = credentails_json_data.get('password')
                store.login_flag = True

                scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
                logs_filename = f'Logs/Logs {scrape_time}.txt'
                Luxottica_Scraper(DEBUG, json_results_filename, logs_filename, chrome_path).controller(store, brands)
                
                print('\n\n Downloading Images...')
                data = read_data_from_json_file(DEBUG, json_results_filename, logs_filename)
                
                saving_picture_in_excel(data, excel_results_filename)

            else: print(f'No {credentails_filename} file found')
        else: print(f'No brands found in {start_json_filename} file')
    else: print(f'No {start_json_filename} file found')
    
except Exception as e: print('Exception: '+str(e))
